#!/usr/bin/env python3
"""Estructura Excel completa: hojas, columnas, dashboard MENU, validaciones."""
import re
import shutil
import zipfile
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parents[1]
SRC = ROOT / "REGISTRO FACTURA LISTO.xlsm"
DST = ROOT / "REGISTRO FACTURA LISTO.xlsm"

COLOR_HEADER = "2D3436"
COLOR_HEADER_FONT = "FFFFFF"
COLOR_SECCION = "0984E3"
COLOR_ACENTO = "00B894"

MAESTRA_HEADERS = [
    "TIPO", "NUMERO", "RIF", "PROVEEDOR", "FECHA", "MONEDA",
    "TOTAL BS", "TOTAL USD", "TASA REGISTRO", "NETO BS", "NETO USD",
    "IVA 16%", "RET IVA", "RET ISLR", "PAGADO BS", "PAGADO USD",
    "SALDO BS", "SALDO USD", "DIF CAMBIARIA USD", "DIAS VENCIDA",
    "RECIBIDO FISICO", "RETENCION ENVIADA", "REGISTRADO", "PAGADO", "PARCIAL", "ESTADO",
]


def fix_islr_formulas(xlsm_path: Path) -> None:
    table_path = "xl/tables/table2.xml"
    with zipfile.ZipFile(xlsm_path, "r") as zin:
        data = zin.read(table_path).decode("utf-8")
    for col in ("BASE PARA ISLR", "RETENCION ISLR"):
        data = re.sub(
            rf'(<tableColumn[^>]*name="{re.escape(col)}"[^>]*?)<calculatedColumnFormula>.*?</calculatedColumnFormula>',
            r"\1",
            data,
            flags=re.DOTALL,
        )
    tmp = xlsm_path.with_suffix(".tmp.zip")
    with zipfile.ZipFile(xlsm_path, "r") as zin:
        with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                payload = zin.read(item.filename)
                if item.filename == table_path:
                    payload = data.encode("utf-8")
                zout.writestr(item, payload)
    shutil.move(str(tmp), str(xlsm_path))


def add_table_columns(ws, table_name: str, new_names: list) -> None:
    if table_name not in ws.tables:
        return
    tbl = ws.tables[table_name]
    min_col, min_row, max_col, max_row = openpyxl.utils.range_boundaries(tbl.ref)
    headers = [ws.cell(min_row, c).value for c in range(min_col, max_col + 1)]
    insert_at = max_col + 1
    for name in new_names:
        if name in headers:
            continue
        ws.cell(min_row, insert_at, value=name)
        insert_at += 1
    tbl.ref = (
        f"{openpyxl.utils.get_column_letter(min_col)}{min_row}:"
        f"{openpyxl.utils.get_column_letter(insert_at - 1)}{max_row}"
    )


def setup_auditoria(wb: openpyxl.Workbook) -> None:
    if "AUDITORIA" in wb.sheetnames:
        del wb["AUDITORIA"]
    ws = wb.create_sheet("AUDITORIA")
    ws.sheet_properties.tabColor = "636E72"
    headers = ["FECHA", "USUARIO", "ACCION", "DETALLE"]
    for i, h in enumerate(headers, start=2):
        ws.cell(3, i, value=h)
    ref = f"B3:E503"
    tab = Table(displayName="AUDITORIA_LOG", ref=ref)
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium4", showRowStripes=True)
    ws.add_table(tab)
    ws["B2"] = "LOG DE AUDITORIA"
    ws["B2"].font = Font(bold=True, size=12, color="FFFFFF")
    ws["B2"].fill = PatternFill("solid", fgColor=COLOR_HEADER)


def setup_resumen(wb: openpyxl.Workbook) -> None:
    if "RESUMEN PROVEEDOR" in wb.sheetnames:
        del wb["RESUMEN PROVEEDOR"]
    ws = wb.create_sheet("RESUMEN PROVEEDOR")
    ws.sheet_properties.tabColor = COLOR_SECCION
    ws["B2"] = "RESUMEN POR PROVEEDOR"
    ws["B2"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["B2"].fill = PatternFill("solid", fgColor=COLOR_SECCION)
    hdrs = ["RIF", "PROVEEDOR", "NETO BS", "PAGADO BS", "SALDO BS", "RET IVA", "RET ISLR", "IVA", "DIF CAMBIARIA USD"]
    for i, h in enumerate(hdrs, start=2):
        c = ws.cell(5, i, value=h)
        c.font = Font(bold=True, color=COLOR_HEADER_FONT)
        c.fill = PatternFill("solid", fgColor=COLOR_HEADER)


def setup_maestra(wb: openpyxl.Workbook) -> None:
    if "BD MAESTRA" in wb.sheetnames:
        ws = wb["BD MAESTRA"]
        # agregar columnas faltantes si la hoja ya existe
        if "BD_MAESTRA" in ws.tables:
            add_table_columns(ws, "BD_MAESTRA", ["DIF CAMBIARIA USD", "DIAS VENCIDA"])
        return
    ws = wb.create_sheet("BD MAESTRA", 1)
    ws.sheet_properties.tabColor = COLOR_ACENTO
    ws.merge_cells("B2:F2")
    ws["B2"] = "CONTROL INTERNO — CHECKLIST DE FACTURAS"
    ws["B2"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["B2"].fill = PatternFill("solid", fgColor=COLOR_SECCION)
    start = 5
    for i, h in enumerate(MAESTRA_HEADERS, start=2):
        ws.cell(start, i, value=h)
    last_col = len(MAESTRA_HEADERS) + 1
    ref = f"B{start}:{openpyxl.utils.get_column_letter(last_col)}{start + 500}"
    tab = Table(displayName="BD_MAESTRA", ref=ref)
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(tab)
    dv = DataValidation(type="list", formula1='"Si,No,Pendiente"', allow_blank=True)
    ws.add_data_validation(dv)
    # RECIBIDO FISICO col 21, RETENCION ENVIADA col 22 (B=2, so col 21 = RECIBIDO)
    col_fisico = openpyxl.utils.get_column_letter(2 + MAESTRA_HEADERS.index("RECIBIDO FISICO"))
    col_ret = openpyxl.utils.get_column_letter(2 + MAESTRA_HEADERS.index("RETENCION ENVIADA"))
    dv.add(f"{col_fisico}6:{col_fisico}505")
    dv2 = DataValidation(type="list", formula1='"Si,No,Pendiente"', allow_blank=True)
    ws.add_data_validation(dv2)
    dv2.add(f"{col_ret}6:{col_ret}505")


def setup_menu_dashboard(wb: openpyxl.Workbook) -> None:
    ws = wb["MENU"]
    fill = PatternFill("solid", fgColor=COLOR_SECCION)
    title_font = Font(bold=True, size=12, color="FFFFFF")
    ws.merge_cells("B6:E6")
    ws["B6"] = "PANEL DE CONTROL"
    ws["B6"].font = title_font
    ws["B6"].fill = fill
    labels = [
        ("B8", "Saldo pendiente (Bs):"),
        ("B9", "Saldo pendiente (USD):"),
        ("B10", "Sin recibir físico:"),
        ("B11", "Sin retención enviada:"),
        ("B12", "Facturas vencidas:"),
    ]
    for cell, txt in labels:
        ws[cell] = txt
        ws[cell].font = Font(bold=True, size=10)
    ws["D8"].number_format = "#,##0.00"
    ws["D9"].number_format = "#,##0.00"
    ws["B3"] = "Dashboard se actualiza al abrir y al guardar facturas/pagos."
    ws["B3"].font = Font(size=9, color="636E72")


def main():
    work = SRC
    if not work.exists():
        work = ROOT / "REGISTRO FACTURA DEFINITIVO.xlsm"
    fix_islr_formulas(work)
    wb = openpyxl.load_workbook(work, keep_vba=True)
    add_table_columns(wb["REGISTRO FACT"], "REGISTROFACTURAS", ["MONEDA", "TASA REGISTRO", "TOTAL USD", "DETALLE ISLR"])
    add_table_columns(wb["BD PAGOS"], "BDPAGOS", ["ANTICIPO APLICADO"])
    setup_maestra(wb)
    setup_auditoria(wb)
    setup_resumen(wb)
    setup_menu_dashboard(wb)
    wb.save(work)
    print(f"Estructura actualizada: {work}")


if __name__ == "__main__":
    main()
