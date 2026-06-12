#!/usr/bin/env python3
"""
Actualiza la estructura del xlsm: corrige tablas XML y agrega hoja BD MAESTRA.
Conserva el proyecto VBA existente (keep_vba=True).
"""
import copy
import re
import shutil
import zipfile
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = Path(__file__).resolve().parents[1]
SRC = ROOT / "REGISTRO FACTURA DEFINITIVO.xlsm"
DST = ROOT / "REGISTRO FACTURA DEFINITIVO_CORREGIDO.xlsm"

# Paleta "no albatros" — teal, coral, menta (sin grises planos)
COLOR_HEADER = "2D3436"
COLOR_HEADER_FONT = "FFFFFF"
COLOR_ACENTO = "00B894"
COLOR_SECCION = "0984E3"

MAESTRA_HEADERS = [
    "TIPO",
    "NUMERO",
    "RIF",
    "PROVEEDOR",
    "FECHA",
    "MONEDA",
    "TOTAL BS",
    "TOTAL USD",
    "TASA REGISTRO",
    "NETO BS",
    "NETO USD",
    "IVA 16%",
    "RET IVA",
    "RET ISLR",
    "PAGADO BS",
    "PAGADO USD",
    "SALDO BS",
    "SALDO USD",
    "RECIBIDO FISICO",
    "RETENCION ENVIADA",
    "REGISTRADO",
    "PAGADO",
    "PARCIAL",
    "ESTADO",
]


def fix_registrofacturas_table_xml(xlsm_path: Path) -> None:
    """Elimina fórmulas #REF! en columnas que VBA llena manualmente (multi-ISLR)."""
    table_path = "xl/tables/table2.xml"
    with zipfile.ZipFile(xlsm_path, "r") as zin:
        data = zin.read(table_path).decode("utf-8")

    # Quitar calculatedColumnFormula de BASE PARA ISLR y RETENCION ISLR
    for col_name in ("BASE PARA ISLR", "RETENCION ISLR"):
        pattern = (
            rf'(<tableColumn[^>]*name="{re.escape(col_name)}"[^>]*?)'
            r'<calculatedColumnFormula>.*?</calculatedColumnFormula>'
        )
        data = re.sub(pattern, r"\1", data, flags=re.DOTALL)

    tmp = xlsm_path.with_suffix(".tmp.zip")
    with zipfile.ZipFile(xlsm_path, "r") as zin:
        with zipfile.ZipFile(tmp, "w", compression=zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                payload = zin.read(item.filename)
                if item.filename == table_path:
                    payload = data.encode("utf-8")
                zout.writestr(item, payload)
    shutil.move(str(tmp), str(xlsm_path))


def style_header_row(ws, row: int, col_count: int) -> None:
    fill = PatternFill("solid", fgColor=COLOR_HEADER)
    font = Font(bold=True, color=COLOR_HEADER_FONT, name="Segoe UI", size=10)
    for c in range(1, col_count + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")


def add_bd_maestra(wb: openpyxl.Workbook) -> None:
    if "BD MAESTRA" in wb.sheetnames:
        del wb["BD MAESTRA"]

    ws = wb.create_sheet("BD MAESTRA", 1)  # después de MENU
    ws.sheet_properties.tabColor = COLOR_ACENTO

    title_fill = PatternFill("solid", fgColor=COLOR_SECCION)
    ws.merge_cells("B2:E2")
    t = ws["B2"]
    t.value = "CONTROL INTERNO — CHECKLIST DE FACTURAS"
    t.font = Font(bold=True, size=14, color="FFFFFF", name="Segoe UI")
    t.fill = title_fill
    t.alignment = Alignment(horizontal="left", vertical="center")

    start_row = 5
    for i, h in enumerate(MAESTRA_HEADERS, start=1):
        ws.cell(row=start_row, column=i + 1, value=h)

    style_header_row(ws, start_row, len(MAESTRA_HEADERS))

    last_col = len(MAESTRA_HEADERS) + 1
    last_row = start_row + 500
    ref = f"B{start_row}:{openpyxl.utils.get_column_letter(last_col)}{last_row}"

    tab = Table(displayName="BD_MAESTRA", ref=ref)
    tab.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(tab)

    widths = {
        2: 8, 3: 14, 4: 14, 5: 22, 6: 12, 7: 10,
        8: 14, 9: 14, 10: 12, 11: 14, 12: 14,
        13: 12, 14: 12, 15: 12, 16: 12, 17: 14,
        18: 14, 19: 14, 20: 14, 21: 16, 22: 16,
        23: 12, 24: 12, 25: 12, 26: 18,
    }
    for col, w in widths.items():
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    note = ws["B3"]
    note.value = (
        "Filtros automáticos activos. "
        "RECIBIDO FISICO / RETENCION ENVIADA: edite Si/No/Pendiente. "
        "REGISTRADO/PAGADO/PARCIAL se actualizan al guardar facturas y pagos."
    )
    note.font = Font(size=9, color="636E72", name="Segoe UI")


def add_registrofacturas_columns(wb: openpyxl.Workbook) -> None:
    ws = wb["REGISTRO FACT"]
    if "REGISTROFACTURAS" not in ws.tables:
        return

    tbl = ws.tables["REGISTROFACTURAS"]
    min_col, min_row, max_col, max_row = openpyxl.utils.range_boundaries(tbl.ref)
    header_row = min_row

    headers = [ws.cell(header_row, c).value for c in range(min_col, max_col + 1)]
    new_cols = ["MONEDA", "TASA REGISTRO", "TOTAL USD"]
    insert_at = max_col + 1

    for name in new_cols:
        if name in headers:
            continue
        ws.cell(header_row, insert_at, value=name)
        insert_at += 1

    # Extender ref de la tabla
    new_max_col = insert_at - 1
    tbl.ref = (
        f"{openpyxl.utils.get_column_letter(min_col)}{min_row}:"
        f"{openpyxl.utils.get_column_letter(new_max_col)}{max_row}"
    )


def main() -> None:
    if not SRC.exists():
        raise SystemExit(f"No se encontró: {SRC}")

    shutil.copy2(SRC, DST)
    fix_registrofacturas_table_xml(DST)

    wb = openpyxl.load_workbook(DST, keep_vba=True)
    add_registrofacturas_columns(wb)
    add_bd_maestra(wb)
    wb.save(DST)
    print(f"Archivo generado: {DST}")


if __name__ == "__main__":
    main()
