#!/usr/bin/env python3
"""
Importa todos los datos de REGISTRO FACTURA LISTO.xlsm a SQLite (backend/prisma/dev.db).
"""
from __future__ import annotations

import json
import re
import sqlite3
import uuid
from datetime import datetime
from pathlib import Path

import openpyxl
import openpyxl.utils

ROOT = Path(__file__).resolve().parents[1]
XLSM = ROOT / "REGISTRO FACTURA LISTO.xlsm"
DB = ROOT / "backend" / "prisma" / "dev.db"


def cid() -> str:
    return f"c{uuid.uuid4().hex[:24]}"


def s(v) -> str | None:
    if v is None:
        return None
    t = str(v).strip()
    if not t or t in ("-", "0", "None"):
        return None
    return t


def is_t2(v) -> bool:
    return str(v).strip().upper() == "T2"


def num(v) -> float | None:
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip().replace(",", ".")
        if not v or v in ("-", "T2"):
            return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def parse_date(v) -> datetime | None:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v
    t = str(v).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(t, fmt)
        except ValueError:
            continue
    return None


def iso_date(d: datetime) -> str:
    return d.strftime("%Y-%m-%dT%H:%M:%S.000Z")


def read_table(ws, table_name: str) -> tuple[list[str], list[list]]:
    if table_name not in ws.tables:
        return [], []
    tbl = ws.tables[table_name]
    min_col, min_row, max_col, max_row = openpyxl.utils.range_boundaries(tbl.ref)
    headers = [ws.cell(min_row, c).value for c in range(min_col, max_col + 1)]
    rows = []
    for r in range(min_row + 1, max_row + 1):
        row = [ws.cell(r, c).value for c in range(min_col, max_col + 1)]
        if any(v is not None and str(v).strip() != "" for v in row):
            rows.append(row)
    return headers, rows


def row_dict(headers: list, row: list) -> dict:
    return {str(h): row[i] if i < len(row) else None for i, h in enumerate(headers) if h}


def calc_iva_neto(
    total_bs: float,
    exento_bs: float,
    retencion_iva_pct: str,
    retencion_islr: float,
    tasa: float | None,
    moneda: str,
) -> dict:
    grabado = max(0.0, total_bs - exento_bs)
    base_imponible = grabado / 1.16
    iva16 = grabado - base_imponible
    r = (retencion_iva_pct or "100%").upper()
    ret_iva_pct = 0.75 if "75" in r else (0.0 if "EXENT" in r else 1.0)
    retencion_iva = iva16 * ret_iva_pct
    monto_apagar = total_bs - retencion_iva - retencion_islr
    monto_usd = None
    if tasa and tasa > 0:
        monto_usd = round(monto_apagar / tasa, 2)
    return {
        "grabadoBs": round(grabado, 2),
        "baseImponible": round(base_imponible, 2),
        "iva16": round(iva16, 2),
        "retencionIva": round(retencion_iva, 2),
        "montoAPagar": round(monto_apagar, 2),
        "montoAPagarUsd": monto_usd,
    }


def clear_db(conn: sqlite3.Connection) -> None:
    conn.executescript(
        """
        DELETE FROM Pago;
        DELETE FROM Factura;
        DELETE FROM Proveedor;
        DELETE FROM Tasa;
        DELETE FROM TabIslr;
        DELETE FROM ConfigItem;
        DELETE FROM AuditoriaLog;
        """
    )


def import_tab_islr(conn: sqlite3.Connection, wb) -> int:
    ws = wb["TAB ISLR"]
    headers, rows = read_table(ws, "TAB_ISLR")
    n = 0
    for orden, row in enumerate(rows):
        d = row_dict(headers, row)
        concepto = s(d.get("CONCEPTO DEL PAGO"))
        if not concepto:
            continue
        pnr_raw = d.get("PNR")
        pjd_raw = d.get("PJD")
        pjnd_raw = d.get("PJND")
        pnnr_raw = d.get("PNNR")
        conn.execute(
            """
            INSERT INTO TabIslr (id, concepto, basePnr, pnr, pagosMinBs, sustraendoBs,
                basePjd, pjd, basePjnd, pjnd, basePnnr, pnnr,
                t2Pnr, t2Pjd, t2Pjnd, t2Pnnr, orden)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """,
            (
                cid(),
                concepto,
                num(d.get("BASE PNR")),
                num(pnr_raw) if not is_t2(pnr_raw) else None,
                num(d.get("PAGOS > A BS.")),
                num(d.get("SUSTR EN BS.")),
                num(d.get("BASE PJD")),
                num(pjd_raw) if not is_t2(pjd_raw) else None,
                num(d.get("BASE PJND")),
                num(pjnd_raw) if not is_t2(pjnd_raw) else None,
                num(d.get("BASE PNNR")),
                num(pnnr_raw) if not is_t2(pnnr_raw) else None,
                1 if is_t2(pnr_raw) else 0,
                1 if is_t2(pjd_raw) else 0,
                1 if is_t2(pjnd_raw) else 0,
                1 if is_t2(pnnr_raw) else 0,
                orden,
            ),
        )
        n += 1
    return n


def import_config(conn: sqlite3.Connection, wb) -> int:
    ws = wb["TABLAS"]
    n = 0
    mapping = [
        ("CAUSADO", "causado", None),
        ("ESTACIONES", "estacion", None),
        ("BANCOS", "banco", "COD"),
        ("IVA", "iva", None),
        ("BCODEPAGOS", "banco_pago", None),
    ]
    for table_name, cat, extra_col in mapping:
        headers, rows = read_table(ws, table_name)
        if not headers:
            continue
        main_h = headers[0]
        extra_idx = headers.index(extra_col) if extra_col and extra_col in headers else None
        for orden, row in enumerate(rows):
            val = s(row[0])
            if not val:
                continue
            extra = s(row[extra_idx]) if extra_idx is not None and extra_idx < len(row) else None
            conn.execute(
                "INSERT INTO ConfigItem (id, categoria, valor, extra, orden) VALUES (?,?,?,?,?)",
                (cid(), cat, val, extra, orden),
            )
            n += 1
    return n


def import_tasas(conn: sqlite3.Connection, wb) -> int:
    ws = wb["TASAS"]
    headers, rows = read_table(ws, "TASAS2026")
    n = 0
    for row in rows:
        d = row_dict(headers, row)
        fecha = parse_date(d.get("FECHA"))
        valor = num(d.get("TASA ") or d.get("TASA"))
        if not fecha or not valor:
            continue
        conn.execute(
            "INSERT INTO Tasa (id, fecha, valor) VALUES (?,?,?)",
            (cid(), iso_date(fecha), valor),
        )
        n += 1
    return n


def import_proveedores(conn: sqlite3.Connection, wb) -> dict[str, str]:
    ws = wb["BD PROVEEDORES"]
    headers, rows = read_table(ws, "BD_Proveedores")
    prov_map: dict[str, str] = {}
    now = iso_date(datetime.now())
    for row in rows:
        d = row_dict(headers, row)
        rif = s(d.get("ID"))
        nombre = s(d.get("NOMBRE"))
        if not rif or not nombre:
            continue
        pid = cid()
        prov_map[rif] = pid
        ret = s(d.get("RETENCION")) or "100%"
        conn.execute(
            """
            INSERT INTO Proveedor (id, rif, nombre, tipoIslr, direccion, telefono, email,
                numCuenta, banco, titular, idTitular, retencionIva, estacion, referido,
                servicio, notas, createdAt, updatedAt)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """,
            (
                pid,
                rif,
                nombre,
                s(d.get("TIPO")) or "PNR",
                s(d.get("DIRECCION")),
                s(d.get("TELEFONO")),
                s(d.get("EMAILS")),
                s(d.get("NUM CUENTA RPOVEEDOR")),
                s(d.get("BANCO")),
                s(d.get("TITULAR DE LA CTA")),
                s(d.get("ID TITULAR")),
                ret,
                s(d.get("ESTACION")),
                s(d.get("REFERIDO POR")),
                s(d.get("SERVICIO")),
                s(d.get("NOTAS")),
                now,
                now,
            ),
        )
    return prov_map


def import_facturas(
    conn: sqlite3.Connection,
    wb,
    prov_by_rif: dict[str, str],
    prov_names: dict[str, str],
    prov_ret_iva: dict[str, str],
) -> dict[str, str]:
    ws = wb["REGISTRO FACT"]
    headers, rows = read_table(ws, "REGISTROFACTURAS")
    maestra_ws = wb["BD MAESTRA"]
    maestra_headers, maestra_rows = read_table(maestra_ws, "BD_MAESTRA")
    maestra_idx = {h: i for i, h in enumerate(maestra_headers) if h}
    maestra_map: dict[tuple, dict] = {}
    for mr in maestra_rows:
        if not maestra_headers:
            break
        md = row_dict(maestra_headers, mr)
        t, n, r = s(md.get("TIPO")), s(md.get("NUMERO")), s(md.get("RIF"))
        if t and n and r:
            maestra_map[(t, n, r)] = md

    factura_ids: dict[tuple, str] = {}
    now = iso_date(datetime.now())
    n = 0

    for row in rows:
        d = row_dict(headers, row)
        tipo = s(d.get("TIPO"))
        numero = s(d.get("NUMERO"))
        rif = s(d.get("RIF"))
        if not tipo or not numero or not rif:
            continue

        key = (tipo, numero, rif)
        if key in factura_ids:
            continue

        fecha = parse_date(d.get("FECHA")) or datetime.now()
        proveedor_nombre = s(d.get("PROVEEDOR")) or prov_names.get(rif) or rif
        exento = num(d.get("EXENTO")) or 0.0
        total_bs = num(d.get("TOTAL FACTURA")) or 0.0
        moneda = s(d.get("MONEDA")) or "Bs"
        tasa = num(d.get("TASA REGISTRO")) or num(d.get("TASA DEL DIA"))
        total_usd = num(d.get("TOTAL USD"))
        if moneda == "USD" and total_usd and tasa:
            total_bs = total_usd * tasa
        elif total_bs and tasa and not total_usd:
            total_usd = round(total_bs / tasa, 2)

        ret_islr = num(d.get("RETENCION ISLR")) or 0.0
        ret_iva_str = prov_ret_iva.get(rif, "100%")
        calc = calc_iva_neto(total_bs, exento, ret_iva_str, ret_islr, tasa, moneda)

        grabado = num(d.get("GRABADO")) or calc["grabadoBs"]
        base_imp = num(d.get("BASE IMPONIBLE")) or calc["baseImponible"]
        iva16 = num(d.get("IVA 16%")) or calc["iva16"]
        ret_iva = num(d.get("RETENCION IVA")) or calc["retencionIva"]
        monto = num(d.get("MONTO A PAGAR")) or calc["montoAPagar"]
        monto_usd = num(d.get("MONTO A PAGAR $")) or calc["montoAPagarUsd"]
        base_islr = num(d.get("BASE PARA ISLR")) or 0.0

        md = maestra_map.get(key, {})
        rec_fisico = s(md.get("RECIBIDO FISICO")) or "Pendiente"
        ret_env = s(md.get("RETENCION ENVIADA")) or "Pendiente"
        if rec_fisico.lower() in ("si", "sí"):
            rec_fisico = "Sí"
        if ret_env.lower() in ("si", "sí"):
            ret_env = "Sí"

        detalle = s(d.get("DETALLE ISLR"))
        if detalle and not detalle.startswith("["):
            detalle = json.dumps([{"concepto": detalle, "monto": base_islr or total_bs}])

        fid = cid()
        factura_ids[key] = fid
        prov_id = prov_by_rif.get(rif)

        conn.execute(
            """
            INSERT INTO Factura (
                id, tipo, numero, rif, proveedorId, proveedorNombre, fecha, causado, concepto,
                diasCredito, moneda, tasaRegistro, totalBs, totalUsd, exentoBs, descripcionIslr,
                detalleIslr, baseIslr, retencionIslr, grabadoBs, baseImponible, iva16,
                retencionIva, montoAPagar, montoAPagarUsd, recibidoFisico, retencionEnviada,
                createdAt, updatedAt
            ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """,
            (
                fid,
                tipo,
                numero,
                rif,
                prov_id,
                proveedor_nombre,
                iso_date(fecha),
                s(d.get("CAUSADO")),
                s(d.get("CONCEPTO")),
                int(num(d.get("DIAS CREDITO")) or 0),
                moneda,
                tasa,
                total_bs,
                total_usd,
                exento,
                s(d.get("DESCRIPCION ISLR")),
                detalle,
                base_islr,
                ret_islr,
                grabado,
                base_imp,
                iva16,
                ret_iva,
                monto,
                monto_usd,
                rec_fisico,
                ret_env,
                now,
                now,
            ),
        )
        n += 1

    return factura_ids


def import_pagos(conn: sqlite3.Connection, wb, factura_ids: dict[tuple, str]) -> int:
    ws = wb["BD PAGOS"]
    headers, rows = read_table(ws, "BDPAGOS")
    n = 0
    for row in rows:
        d = row_dict(headers, row)
        fecha = parse_date(d.get("FECHA"))
        rif = s(d.get("RIF"))
        if not fecha or not rif:
            continue
        doc = s(d.get("DOCUMENTO")) or ""
        factura_id = None
        for key, fid in factura_ids.items():
            t, fact_num, r = key
            if r == rif and (doc == f"{t}-{fact_num}" or doc == fact_num or doc in f"{t}-{fact_num}"):
                factura_id = fid
                break

        conn.execute(
            """
            INSERT INTO Pago (id, fecha, rif, proveedor, documento, banco, referencia,
                pagadoBs, pagadoUsd, tasa, observacion, estadoAnticipo, anticipoAplicado,
                facturaId, createdAt)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """,
            (
                cid(),
                iso_date(fecha),
                rif,
                s(d.get("PROVEEDOR")) or "",
                doc,
                s(d.get("BANCO")) or "",
                s(d.get("REFERENCIA")) or "",
                num(d.get("PAGADO Bs.")),
                num(d.get("PAGADO USD")),
                num(d.get("TASA")),
                s(d.get("OBSERVACION")),
                s(d.get("ESTADO ANTICIPO")),
                num(d.get("ANTICIPO APLICADO")),
                factura_id,
                iso_date(datetime.now()),
            ),
        )
        n += 1
    return n


def import_auditoria(conn: sqlite3.Connection, wb) -> int:
    if "AUDITORIA" not in wb.sheetnames:
        return 0
    ws = wb["AUDITORIA"]
    headers, rows = read_table(ws, "AUDITORIA_LOG")
    n = 0
    for row in rows:
        d = row_dict(headers, row)
        accion = s(d.get("ACCION"))
        if not accion:
            continue
        fecha = parse_date(d.get("FECHA")) or datetime.now()
        conn.execute(
            "INSERT INTO AuditoriaLog (id, fecha, usuario, accion, detalle) VALUES (?,?,?,?,?)",
            (cid(), iso_date(fecha), s(d.get("USUARIO")), accion, s(d.get("DETALLE"))),
        )
        n += 1
    return n


def main() -> None:
    if not XLSM.exists():
        raise SystemExit(f"No se encontró {XLSM}")

    print(f"Leyendo {XLSM.name}...")
    wb = openpyxl.load_workbook(XLSM, keep_vba=True, data_only=True)

    conn = sqlite3.connect(DB)
    try:
        clear_db(conn)
        tab_n = import_tab_islr(conn, wb)
        cfg_n = import_config(conn, wb)
        tasa_n = 0  # Tasas: histórico BCV automático (npm run db:bcv-rebuild)

        prov_ids = import_proveedores(conn, wb)
        prov_names: dict[str, str] = {}
        prov_ret: dict[str, str] = {}
        for row in conn.execute("SELECT rif, nombre, retencionIva FROM Proveedor"):
            prov_names[row[0]] = row[1]
            prov_ret[row[0]] = row[2]

        fact_ids = import_facturas(conn, wb, prov_ids, prov_names, prov_ret)
        pago_n = import_pagos(conn, wb, fact_ids)
        aud_n = import_auditoria(conn, wb)

        conn.commit()
        print("Importación completada:")
        print(f"  Tab ISLR:     {tab_n}")
        print(f"  Config:       {cfg_n}")
        print(f"  Tasas:        {tasa_n}")
        print(f"  Proveedores:  {len(prov_ids)}")
        print(f"  Facturas:     {len(fact_ids)}")
        print(f"  Pagos:        {pago_n}")
        print(f"  Auditoría:    {aud_n}")
        print(f"  Base de datos: {DB}")
    finally:
        conn.close()


if __name__ == "__main__":
    main()
